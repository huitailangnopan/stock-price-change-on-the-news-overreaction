import re
import os
import pandas as pd
from datetime import datetime
import csv
import openpyxl
from func import *
def match(ticker):
    os.chdir("D:\\high-frequency data\\analysis")
    wb = openpyxl.Workbook()
    wb.save(filename=ticker + '.xlsx')
    os.chdir("D:\\high-frequency data\\news")
    wb = openpyxl.load_workbook(ticker + '.xlsx')
    ws = wb.active
    os.chdir("D:\\high-frequency data\\analysis")
    ticker = ticker.upper()
    wb = openpyxl.Workbook()
    wb.save(filename=ticker + '.xlsx')
    for x in range(1,ws.max_row+1):

        a = ws.cell(row=x,column=2).value
        ticker = ticker.lower()

        dictionary = {'Jan':'01','Feb':'02','Mar':'03','Apr':'04','May':'05','Jun':'06','Jul':'07','Aug':'08','Sep':'09','Oct':'10',
                      'Nov':'11','Dec':'12'}
        x = re.search(",",a)
        y=x.start()

        week=a[:y]
        day = a[y+2:y+4]
        month = a[y+5:y+8]
        month = dictionary[month]
        year = a[y+9:y+13]
        time = a[y+14:y+22]
        hour = time[:2]
        min = time[3:5]
        second = time[6:8]
        date_time_str = day+'/'+month+'/'+year[2:]+' '+hour+':'+min+':'+second                    #'18/09/19 01:55:19'
        date_time_obj = datetime.strptime(date_time_str, '%d/%m/%y %H:%M:%S')
        if hour[0]=='0':
            hour=hour[1:]
        #print(week,day,month,year,time,hour,min,second)
        date = year+month+day
        if tradeDay(date,ticker)==True:
            time_spot = timespot(date_time_obj)
            try:
                if time_spot == "pre_trade":
                    want = pre_trade(date_time_obj,ticker)
                if time_spot == "pre_market":
                    want = pre_market(date_time_obj,ticker)
                if time_spot == "market":
                    want = market(date_time_obj,ticker)
                if time_spot == "post_market":
                    want = post_market(date_time_obj,ticker)
                if time_spot == "post_trade":
                    want = post_trade(date_time_obj,ticker)
            except FileNotFoundError:
                continue
        else:
            continue
        #df = pd.DataFrame(columns=['time', 'ticker', 'status','pre','open', 'volume', 'one min', 'one hour','close price','post',
                           #    'next day opening','next day close','ten days close','one month close'])
        #df = df.append(want)
        os.chdir("D:\\high-frequency data\\analysis")
        ticker = ticker.upper()
        try:
            want.to_excel('temporary.xlsx', index=False, header=False)
            wn = openpyxl.load_workbook('temporary.xlsx')
            wm = wn.active
            wd = openpyxl.load_workbook(ticker + '.xlsx')
            wf = wd.active
            previous = wf.max_row
            for tt in range(1,wm.max_column+1):
                wf.cell(row=previous+1,column=tt).value = wm.cell(row=wm.max_row,column=tt).value
            wd.save(ticker + '.xlsx')
        except AttributeError:
            continue
        '''
        os.chdir("D:\\high-frequency data\\analysis")
        ticker = ticker.upper()
        wb = openpyxl.load_workbook('temporary.xlsx')
        ws = wb.active
        want.to_excel('temporary.xlsx',startrow=ws.max_row+1,index=False,header=False)
        '''
        '''
        wb = openpyxl.load_workbook(ticker+'.xlsx')
        ws = wb.active
        previous = ws.max_row
        wn = openpyxl.load_workbook('temporary.xlsx')
        wm = wn.active
        if wm.max_column<1:
            continue
        for tt in range(1,wm.max_column+1):
            ws.cell(row=previous+1,column=tt).value = wm.cell(row=1,column=tt).value
        wb.save(ticker.upper() + '.xlsx')
'''


