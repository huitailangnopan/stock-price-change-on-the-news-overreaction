import requests
import os
import openpyxl
import time
import json
os.chdir(r"D:\high-frequency data")
arr = os.listdir()
stockfolder=[]
for file in arr:
   if (file.find('.xlsx') == -1):
       if (file.find('.py') == -1):
            stockfolder.append(file)
for q in range(len(stockfolder)):
    ticker = stockfolder[q].upper()
    os.chdir(r"D:\high-frequency data\news")
    wb = openpyxl.Workbook()
    wb.save(filename=ticker+'.xlsx')
    #&source=Business+Wire,GlobeNewsWire,PRNewswire
    api_key = 'zgo4odj3wpfvabi5uolkcogdu72yrjkalug5z6tu'
    exclude = "4/7+Wall+Street,Benzinga,Seeking+Alpha,The+Motley+Fool,Investors+Business+Daily,Zacks+Investment+Research,Barrons,Market+Watch"
    data_list=['02192021-03192021','01192021-02192021','12192020-01192021','11192020-12192020','10192020-11192020','09192020-10192020',
               '08192020-09192020','07192020-08192020','06192020-07192020','05192020-06192020','04192020-05192020',
               '03192020-04192020','02192020-03192020','01192020-02192020','12192019-01192020','11192019-12192019','10192019-11192019','09192019-10192019',
               '08192019-09192019','07192019-08192019','06192019-07192019','05192019-06192019','04192019-05192019',
               '03192019-04192019']
    for y in range(0,len(data_list)):
        os.chdir(r"D:\high-frequency data\news")
        wb = openpyxl.load_workbook(ticker+'.xlsx')
        ws = wb.active
        date=data_list[y]
        list_url = 'https://stocknewsapi.com/api/v1?tickers='+ticker+'&sourceexclude='+exclude+'&date='+date+'&items=50&token='+api_key
        resp = requests.get(list_url)
        getdata = resp.json()
        #print(getdata)
        print(date)
        refined = getdata['data']
        previous = ws.max_row
        for x in range(len(refined)):
            kk = refined[x]
            a=0
            if previous>50:
                for z in range(previous-50,previous+1):
                    if kk['title']==ws.cell(row=z, column=1).value:
                        a=1
            if a==1:
                continue
            article_url = kk['news_url']
            ws.cell(row=x + 1+previous, column=3).value = article_url
            title = kk['title']
            ws.cell(row=x + 1+previous, column=1).value = title
            published_time = kk['date']
            ws.cell(row=x + 1+previous, column=2).value = published_time
            source_name = kk['source_name']
            ws.cell(row=x + 1+previous, column=4).value = source_name
            sentiment = kk['sentiment']
            ws.cell(row=x + 1+previous, column=5).value = sentiment
            text = kk['text']
            ws.cell(row=x + 1+previous, column=6).value = text
        os.chdir(r"G:\high_frequency_data\news")
        wb.save(ticker+'.xlsx')
