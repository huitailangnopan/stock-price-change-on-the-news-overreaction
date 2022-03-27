import requests
import os
import openpyxl
import time
import json
from match import match

os.chdir(r"D:\high-frequency data\pharmarcy")
arr = os.listdir()
stockfolder=[]
for file in arr:
   if (file.find('.xlsx') == -1):
       if (file.find('.py') == -1):
            stockfolder.append(file)
for i in range(0,len(stockfolder)):
    ticker = stockfolder[i]
    os.chdir("D:\\high-frequency data\\analysis")
    wb = openpyxl.load_workbook(ticker + '.xlsx')
    ws = wb.active
    #analysis = ['pre_increase','daily_increase','post_increase','one_min','one_hour','next_day','next_10days','next_month']
    for x in range(2,ws.max_row+1):
        try:
            if ws.cell(row=x,column=3).value == "pre_trade":
                pre_increase = (float(ws.cell(row=x,column=5).value)-float(ws.cell(row=x,column=4).value))/float(ws.cell(row=x,column=4).value)
                daily_increase = (float(ws.cell(row=x,column=9).value)-float(ws.cell(row=x,column=4).value))/float(ws.cell(row=x,column=4).value)
                post_increase = (float(ws.cell(row=x,column=10).value)-float(ws.cell(row=x,column=4).value))/float(ws.cell(row=x,column=4).value)
                next_day = (float(ws.cell(row=x,column=12).value)-float(ws.cell(row=x,column=4).value))/float(ws.cell(row=x,column=4).value)
                next_10days = (float(ws.cell(row=x,column=13).value)-float(ws.cell(row=x,column=4).value))/float(ws.cell(row=x,column=4).value)
                next_month = (float(ws.cell(row=x,column=14).value)-float(ws.cell(row=x,column=4).value))/float(ws.cell(row=x,column=4).value)
                ws.cell(row=x,column=15).value = pre_increase
                ws.cell(row=x, column=16).value = daily_increase
                ws.cell(row=x, column=17).value = post_increase
                ws.cell(row=x, column=20).value = next_day
                ws.cell(row=x, column=21).value = next_10days
                ws.cell(row=x,column=22).value = next_month
            if ws.cell(row=x,column=3).value == "pre_market":
                pre_increase = (float(ws.cell(row=x,column=7).value)-float(ws.cell(row=x,column=5).value))/float(ws.cell(row=x,column=5).value)
                one_min = (float(ws.cell(row=x,column=7).value)-float(ws.cell(row=x,column=5).value))/float(ws.cell(row=x,column=5).value)
                one_hour = (float(ws.cell(row=x, column=8).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                daily_increase = (float(ws.cell(row=x, column=9).value) - float(
                    ws.cell(row=x, column=5).value)) / float(ws.cell(row=x, column=5).value)
                post_increase = (float(ws.cell(row=x, column=10).value) - float(
                    ws.cell(row=x, column=5).value)) / float(ws.cell(row=x, column=5).value)
                next_day = (float(ws.cell(row=x, column=12).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_10days = (float(ws.cell(row=x, column=13).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_month = (float(ws.cell(row=x, column=14).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                ws.cell(row=x, column=15).value = pre_increase
                ws.cell(row=x, column=16).value = daily_increase
                ws.cell(row=x, column=17).value = post_increase
                ws.cell(row=x, column=20).value = next_day
                ws.cell(row=x, column=21).value = next_10days
                ws.cell(row=x, column=22).value = next_month
                ws.cell(row=x,column=18).value = one_min
                ws.cell(row=x,column=19).value = one_hour
            if ws.cell(row=x,column=3).value is None:
                one_min = (float(ws.cell(row=x, column=7).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                one_hour = (float(ws.cell(row=x, column=8).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                daily_increase = (float(ws.cell(row=x, column=9).value) - float(
                    ws.cell(row=x, column=5).value)) / float(ws.cell(row=x, column=5).value)
                post_increase = (float(ws.cell(row=x, column=10).value) - float(
                    ws.cell(row=x, column=5).value)) / float(ws.cell(row=x, column=5).value)
                next_day = (float(ws.cell(row=x, column=12).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_10days = (float(ws.cell(row=x, column=13).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_month = (float(ws.cell(row=x, column=14).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                ws.cell(row=x, column=16).value = daily_increase
                ws.cell(row=x, column=17).value = post_increase
                ws.cell(row=x, column=20).value = next_day
                ws.cell(row=x, column=21).value = next_10days
                ws.cell(row=x, column=22).value = next_month
                ws.cell(row=x, column=18).value = one_min
                ws.cell(row=x, column=19).value = one_hour
            if ws.cell(row=x,column=3).value == "post_market":
                one_min = (float(ws.cell(row=x, column=7).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                one_hour = (float(ws.cell(row=x, column=8).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                post_increase = (float(ws.cell(row=x, column=10).value) - float(
                    ws.cell(row=x, column=5).value)) / float(ws.cell(row=x, column=5).value)
                next_day = (float(ws.cell(row=x, column=12).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_10days = (float(ws.cell(row=x, column=13).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_month = (float(ws.cell(row=x, column=14).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                ws.cell(row=x, column=17).value = post_increase
                ws.cell(row=x, column=20).value = next_day
                ws.cell(row=x, column=21).value = next_10days
                ws.cell(row=x, column=22).value = next_month
                ws.cell(row=x, column=18).value = one_min
                ws.cell(row=x, column=19).value = one_hour
            if ws.cell(row=x,column=3).value == "post_trade":
                next_day = (float(ws.cell(row=x, column=12).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_10days = (float(ws.cell(row=x, column=13).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                next_month = (float(ws.cell(row=x, column=14).value) - float(ws.cell(row=x, column=5).value)) / float(
                    ws.cell(row=x, column=5).value)
                ws.cell(row=x, column=20).value = next_day
                ws.cell(row=x, column=21).value = next_10days
                ws.cell(row=x, column=22).value = next_month
        except ValueError:
            pass
        except TypeError:
            pass
    wb.save(ticker+'.xlsx')

